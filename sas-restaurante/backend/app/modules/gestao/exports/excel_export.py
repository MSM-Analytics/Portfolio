from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from fastapi.responses import StreamingResponse
from io import BytesIO


def export_excel(
    filename: str,
    sheets: dict
):
    """
    sheets = {
        "Nome da Aba": {
            "headers": [...],
            "rows": [ {..}, {..} ]
        }
    }
    """

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, data in sheets.items():
        ws = wb.create_sheet(title=sheet_name)

        headers = data["headers"]
        rows = data["rows"]

        # Cabeçalho
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Dados
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header))

        # Autofiltro e ajuste de coluna
        ws.auto_filter.ref = ws.dimensions

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
